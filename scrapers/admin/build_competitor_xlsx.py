"""Build the consolidated Phase 2.1 competitor review Excel workbook.

Reads SERP + validator CSVs for all 8 markets, emits a single .xlsx with
three sheets sized for marketing-team review and decision capture.

Usage:
    cd /home/ubuntu/app
    .venv/bin/python scrapers/admin/build_competitor_xlsx.py

One-time setup on EC2 (openpyxl is not in scrapers/requirements.txt since
this is an operator tool, not a runtime dependency):
    .venv/bin/pip install openpyxl

Sheets:
  1. Matrix — wide: 15 competitors × 8 markets, color-coded by confidence
  2. Decisions — long: ~120 rows, each (competitor, market) with algorithm
     verdict + dropdown for marketing to fill in active/planned/withdrawn/skip
  3. Signals — raw SERP + validator detail per row (for the curious)

Output: logs/PHASE_2_1_COMPETITOR_REVIEW.xlsx
"""
from __future__ import annotations

import csv
import os
import sys
from datetime import datetime, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run:", file=sys.stderr)
    print("    .venv/bin/pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_SCRAPERS_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PROJECT_ROOT = os.path.dirname(_SCRAPERS_DIR)
sys.path.insert(0, _SCRAPERS_DIR)

from config import COMPETITORS  # noqa: E402

MARKETS = ["sg", "hk", "tw", "my", "th", "ph", "id", "vn"]
MARKET_NAMES = {
    "sg": "Singapore", "hk": "Hong Kong", "tw": "Taiwan",
    "my": "Malaysia", "th": "Thailand", "ph": "Philippines",
    "id": "Indonesia", "vn": "Vietnam",
}
LOGS_DIR = os.path.join(_PROJECT_ROOT, "logs")
OUT_PATH = os.path.join(LOGS_DIR, "PHASE_2_1_COMPETITOR_REVIEW.xlsx")

# Style palette
FILL_STRONG = PatternFill(start_color="C6E8C6", end_color="C6E8C6", fill_type="solid")  # green
FILL_MEDIUM = PatternFill(start_color="FFF1C2", end_color="FFF1C2", fill_type="solid")  # amber
FILL_WEAK   = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # red
FILL_EMPTY  = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # grey
FILL_HEADER = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")  # dark
FILL_WARN   = PatternFill(start_color="FFD27A", end_color="FFD27A", fill_type="solid")  # orange

FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_BOLD = Font(bold=True)

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read_csv(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    with open(path) as f:
        return list(csv.DictReader(f))


def _load_market(market: str) -> dict[str, dict]:
    """Returns {competitor_id: merged_row} for one market."""
    out: dict[str, dict] = {}
    for row in _read_csv(os.path.join(LOGS_DIR, f"serp_research_{market}.csv")):
        cid = row["competitor_id"]
        out[cid] = {
            "queries_appeared": int(row.get("queries_appeared") or 0),
            "best_rank": int(row["best_rank"]) if row.get("best_rank") else None,
            "avg_rank": float(row["avg_rank"]) if row.get("avg_rank") else None,
            "own_brand_hits": int(row.get("own_brand_hits") or 0),
        }
    for row in _read_csv(os.path.join(LOGS_DIR, f"market_presence_{market}.csv")):
        cid = row["competitor_id"]
        out.setdefault(cid, {})
        out[cid].update({
            "S1_local_url": int(row.get("S1_local_url") or 0),
            "S2_local_lang": int(row.get("S2_local_lang") or 0),
            "S3_payment": int(row.get("S3_payment") or 0),
            "S4_app_store": int(row.get("S4_app_store") or 0),
            "S5_wikifx": int(row.get("S5_wikifx") or 0),
            "signal_count": int(row.get("signal_count") or 0),
            "serp_conf": row.get("serp_conf") or "weak",
            "combined": row.get("combined") or "weak",
        })
    return out


def _conf_fill(combined: str | None) -> PatternFill:
    return {"STRONG": FILL_STRONG, "medium": FILL_MEDIUM, "weak": FILL_WEAK}.get(combined, FILL_EMPTY)


def _is_thin_fetch(entry: dict) -> bool:
    return entry.get("serp_conf") == "STRONG" and entry.get("signal_count", 0) == 0


# ─────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────
def build_matrix_sheet(wb: Workbook, matrix: dict) -> None:
    ws = wb.create_sheet("Matrix", 0)
    ws.freeze_panes = "B2"

    # Header
    headers = ["Competitor"] + [m.upper() for m in MARKETS] + ["Marketing Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Rows
    competitor_ids = [c["id"] for c in COMPETITORS]
    for row_idx, cid in enumerate(competitor_ids, start=2):
        ws.cell(row=row_idx, column=1, value=cid).font = FONT_BOLD
        ws.cell(row=row_idx, column=1).border = BORDER
        for col_idx, m in enumerate(MARKETS, start=2):
            entry = matrix.get(m, {}).get(cid)
            if not entry:
                cell = ws.cell(row=row_idx, column=col_idx, value="—")
                cell.fill = FILL_EMPTY
            else:
                label = entry.get("combined", "weak")
                if _is_thin_fetch(entry):
                    label += " ⚠"
                cell = ws.cell(row=row_idx, column=col_idx, value=label)
                cell.fill = _conf_fill(entry.get("combined"))
                if _is_thin_fetch(entry):
                    cell.fill = FILL_WARN
                if entry.get("combined") == "STRONG":
                    cell.font = FONT_BOLD
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        # Notes column (empty for marketing)
        ws.cell(row=row_idx, column=len(MARKETS) + 2).border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = 16
    for i in range(2, len(MARKETS) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[get_column_letter(len(MARKETS) + 2)].width = 40

    # Legend below matrix
    legend_row = len(competitor_ids) + 4
    ws.cell(row=legend_row, column=1, value="Legend:").font = FONT_BOLD
    legend_items = [
        ("STRONG", FILL_STRONG, "Show by default — SERP strong OR ≥3 validator signals"),
        ("medium", FILL_MEDIUM, "Operator decides — review case-by-case"),
        ("weak", FILL_WEAK, "Hide by default — override only if marketing knows otherwise"),
        ("⚠", FILL_WARN, "Thin-content fetch — STRONG SERP but 0 signals; verify manually"),
        ("—", FILL_EMPTY, "Market not yet researched"),
    ]
    for i, (label, fill, desc) in enumerate(legend_items):
        r = legend_row + 1 + i
        c1 = ws.cell(row=r, column=1, value=label)
        c1.fill = fill
        c1.alignment = Alignment(horizontal="center")
        c1.border = BORDER
        ws.cell(row=r, column=2, value=desc)


def build_decisions_sheet(wb: Workbook, matrix: dict) -> None:
    ws = wb.create_sheet("Decisions", 1)
    ws.freeze_panes = "A2"

    headers = [
        "Market", "Competitor", "Algorithm Verdict", "SERP", "Queries",
        "Best Rank", "S1 URL", "S2 Lang", "S3 Pay", "S4 App", "S5 WikiFX",
        "Decision", "Reasoning", "Approved By",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    competitor_ids = [c["id"] for c in COMPETITORS]
    row_num = 2
    # Sort: market → combined confidence rank → competitor id
    conf_rank = {"STRONG": 0, "medium": 1, "weak": 2, None: 3}
    for m in MARKETS:
        market_data = matrix.get(m, {})
        # Build rows for this market, sorted by confidence then id
        rows = []
        for cid in competitor_ids:
            e = market_data.get(cid, {})
            rows.append((conf_rank.get(e.get("combined")), cid, e))
        rows.sort()
        for _, cid, e in rows:
            ws.cell(row=row_num, column=1, value=MARKET_NAMES[m]).border = BORDER
            ws.cell(row=row_num, column=2, value=cid).font = FONT_BOLD
            ws.cell(row=row_num, column=2).border = BORDER
            verdict_cell = ws.cell(row=row_num, column=3, value=e.get("combined") or "—")
            verdict_cell.fill = _conf_fill(e.get("combined"))
            verdict_cell.alignment = Alignment(horizontal="center")
            verdict_cell.border = BORDER
            ws.cell(row=row_num, column=4, value=e.get("serp_conf") or "—").border = BORDER
            ws.cell(row=row_num, column=5, value=e.get("queries_appeared") or 0).border = BORDER
            ws.cell(row=row_num, column=6, value=e.get("best_rank") or "").border = BORDER
            for col_offset, key in enumerate(["S1_local_url", "S2_local_lang", "S3_payment",
                                              "S4_app_store", "S5_wikifx"]):
                v = e.get(key, "")
                cell = ws.cell(row=row_num, column=7 + col_offset, value=("✓" if v == 1 else ""))
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER
            # Decision (empty — for marketing to fill via dropdown)
            for col in range(12, 15):
                ws.cell(row=row_num, column=col).border = BORDER
            row_num += 1

    # Column widths
    widths = [12, 16, 10, 8, 9, 10, 8, 8, 8, 8, 10, 14, 30, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Dropdown validation on Decision column (L = 12)
    dv = DataValidation(
        type="list",
        formula1='"active,planned,withdrawn,skip,emerging"',
        allow_blank=True,
        showDropDown=False,  # show the arrow
    )
    dv.error = "Pick one: active / planned / withdrawn / skip / emerging"
    dv.prompt = "Marketing decision for this (market, competitor)"
    dv.promptTitle = "Decision"
    ws.add_data_validation(dv)
    dv.add(f"L2:L{row_num - 1}")


def build_signals_sheet(wb: Workbook, matrix: dict) -> None:
    """Raw signals reference sheet — same data as Decisions but flat / sortable."""
    ws = wb.create_sheet("Signals (raw)", 2)
    ws.freeze_panes = "A2"

    headers = [
        "Market", "Competitor", "Combined", "SERP Confidence", "Queries Appeared",
        "Total Appearances", "Best Rank", "Avg Rank", "Own Brand Hits",
        "S1 Local URL", "S2 Local Lang", "S3 Local Payment",
        "S4 App Store", "S5 WikiFX Local", "Signal Count",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    row_num = 2
    for m in MARKETS:
        market_data = matrix.get(m, {})
        for cid in [c["id"] for c in COMPETITORS]:
            e = market_data.get(cid, {})
            vals = [
                MARKET_NAMES[m], cid,
                e.get("combined") or "—",
                e.get("serp_conf") or "weak",
                e.get("queries_appeared") or 0,
                None,  # total_appearances — derived; skip
                e.get("best_rank") or "",
                e.get("avg_rank") or "",
                e.get("own_brand_hits") or 0,
                e.get("S1_local_url", ""),
                e.get("S2_local_lang", ""),
                e.get("S3_payment", ""),
                e.get("S4_app_store", ""),
                e.get("S5_wikifx", ""),
                e.get("signal_count", 0),
            ]
            for col, v in enumerate(vals, start=1):
                ws.cell(row=row_num, column=col, value=v)
            row_num += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16


def build_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("How to use", 3)
    ws.column_dimensions["A"].width = 100
    instructions = [
        ("Phase 2.1 — Competitor Review", FONT_HEADER),
        ("", None),
        ("Goal: decide which competitors should appear by default on each /markets/<code> view.", None),
        ("", None),
        ("Sheets:", FONT_BOLD),
        ("  1. Matrix — quick scan view of all 15 competitors × 8 markets, color-coded.", None),
        ("  2. Decisions — one row per (competitor, market). Fill the Decision column.", None),
        ("  3. Signals (raw) — underlying SERP + validator data for reference.", None),
        ("", None),
        ("Decision values (dropdown on Decisions sheet):", FONT_BOLD),
        ("  active     — competitor IS active in this market; show on the market view", None),
        ("  planned    — broker has announced market entry; show with 'Watching' badge", None),
        ("  withdrawn  — broker withdrew from market; keep historical data, hide from default", None),
        ("  skip       — broker is NOT in this market; hide", None),
        ("  emerging   — strong SERP signal but no operator approval yet; show in 'Emerging' rail", None),
        ("", None),
        ("Workflow:", FONT_BOLD),
        ("  1. Open the Decisions sheet.", None),
        ("  2. For each row, pick a value from the Decision dropdown (column L).", None),
        ("  3. Add notes in Reasoning column where useful.", None),
        ("  4. Initial / sign off in Approved By column.", None),
        ("  5. Save the file and send back to engineering.", None),
        ("", None),
        ("Edge cases to watch:", FONT_BOLD),
        ("  - ⚠ symbol in Matrix = thin-content fetch (likely JS-rendered shell). Algorithm verdict", None),
        ("    is unreliable for these — verify by visiting the broker's site directly.", None),
        ("  - weak verdict from algorithm doesn't always mean inactive — marketing's word-of-mouth", None),
        ("    knowledge wins. Override to active if you know the broker is operating there.", None),
        ("  - STRONG verdict with low validator signals (e.g., 0 or 1) means the broker shows up", None),
        ("    in SERP but doesn't have many operational markers — could be SEO-only, not really", None),
        ("    serving the market. Worth a sanity check.", None),
    ]
    for i, (text, font) in enumerate(instructions, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font


def main() -> int:
    matrix: dict[str, dict[str, dict]] = {}
    found = 0
    for m in MARKETS:
        matrix[m] = _load_market(m)
        if matrix[m]:
            found += 1
            print(f"  {m}: loaded {len(matrix[m])} row(s)")
        else:
            print(f"  {m}: NO DATA (CSVs missing — column will show '—' in matrix)")

    wb = Workbook()
    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_matrix_sheet(wb, matrix)
    build_decisions_sheet(wb, matrix)
    build_signals_sheet(wb, matrix)
    build_readme_sheet(wb)

    os.makedirs(LOGS_DIR, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\n✓ Wrote {OUT_PATH}")
    print(f"  Markets covered: {found}/{len(MARKETS)}")
    print(f"  Sheets: Matrix · Decisions · Signals (raw) · How to use")
    return 0


if __name__ == "__main__":
    sys.exit(main())
