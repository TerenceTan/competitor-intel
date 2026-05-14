"""Operator-side seed import: read PHASE_2_1_COMPETITOR_REVIEW.xlsx → competitor_markets.

Reads the marketing-approved Decisions sheet from the consolidated review
workbook and upserts rows into the competitor_markets table (Plan 02.1-01).

Usage:
    cd /home/ubuntu/app
    .venv/bin/python scrapers/admin/import_market_decisions.py
    .venv/bin/python scrapers/admin/import_market_decisions.py --purge        # destructive re-import
    .venv/bin/python scrapers/admin/import_market_decisions.py --input <path> # alternate workbook
    .venv/bin/python scrapers/admin/import_market_decisions.py --dry-run      # print plan, no DB writes

One-time setup on EC2 (openpyxl is not in scrapers/requirements.txt since
this is an operator tool, not a runtime dependency):
    .venv/bin/pip install openpyxl

Semantics (per D2.1-09):
  - INSERT OR REPLACE on (competitor_id, market_code) — upsert.
  - Blank Decision cells are SKIPPED (marketing hasn't decided yet).
  - Decision='skip' values are SKIPPED (marketing decided "not relevant").
  - Valid Decisions (active | planned | withdrawn | emerging) are UPSERTed.
  - WITHOUT --purge: removing a row from the spreadsheet does NOT remove
    it from the DB. Upsert-only.
  - WITH --purge: for every market_code that appears in the spreadsheet,
    all existing rows for that market are DELETEd BEFORE inserting. This
    is the "spreadsheet is the source of truth, mirror it exactly" mode.
    Use with care.

Audit trail (per D2.1-10): every successful run writes one row to
change_events with field_name='market_curation_imported' and a JSON
summary in new_value (rows_inserted, rows_skipped_blank, rows_skipped_skip,
purge_mode). Failures (e.g. missing workbook) write no audit row.

Mirror of operator-side workbook structure:
  - Sheet name: 'Decisions'
  - Header row: row 1 (frozen)
  - Data rows: row 2..N
  - Column mapping (1-indexed):
      1=Market (e.g. 'Singapore'), 2=Competitor (id),
      12=Decision (dropdown), 13=Reasoning, 14=Approved By
  - Source: scrapers/admin/build_competitor_xlsx.py
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from contextlib import closing
from datetime import datetime, timezone

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run:", file=sys.stderr)
    print("    .venv/bin/pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_SCRAPERS_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PROJECT_ROOT = os.path.dirname(_SCRAPERS_DIR)
sys.path.insert(0, _SCRAPERS_DIR)

from db_utils import get_db  # noqa: E402

# Mirror of scrapers/admin/build_competitor_xlsx.py:46-50. Keep in sync if
# either file changes (both reference the canonical APAC v1 8-market list).
MARKETS = ["sg", "hk", "tw", "my", "th", "ph", "id", "vn"]
MARKET_NAMES = {
    "sg": "Singapore", "hk": "Hong Kong", "tw": "Taiwan",
    "my": "Malaysia", "th": "Thailand", "ph": "Philippines",
    "id": "Indonesia", "vn": "Vietnam",
}
NAME_TO_CODE = {v: k for k, v in MARKET_NAMES.items()}

# Valid statuses for competitor_markets per Plan 02.1-01's CHECK constraint.
# Note: 'skip' is in the xlsx dropdown but is NOT a valid status — it means
# "don't insert this row at all". Treated the same as a blank Decision.
VALID_STATUSES = {"active", "planned", "withdrawn", "emerging"}

DEFAULT_XLSX_PATH = os.path.join(_PROJECT_ROOT, "logs", "PHASE_2_1_COMPETITOR_REVIEW.xlsx")
# Audit-row constants (D2.1-10). competitor_id='pepperstone' is guaranteed to
# exist because db_utils.get_db() runs an INSERT OR IGNORE for the self-benchmark
# row on every connection open — so the FK to competitors(id) is always satisfied
# without introducing a new sentinel ID.
AUDIT_DOMAIN = "admin"
AUDIT_FIELD = "market_curation_imported"
AUDIT_COMPETITOR_ID = "pepperstone"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Import marketing-approved per-market curation decisions."
    )
    p.add_argument(
        "--input",
        default=DEFAULT_XLSX_PATH,
        help=f"Path to the Decisions xlsx (default: {DEFAULT_XLSX_PATH})",
    )
    p.add_argument(
        "--purge",
        action="store_true",
        help=(
            "DESTRUCTIVE: DELETE all rows for any market_code appearing in the "
            "spreadsheet before inserting. Use only when the spreadsheet is "
            "intended to be the exact mirror of the DB state."
        ),
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Print the planned actions and the summary but make no DB writes.",
    )
    return p.parse_args()


def parse_decisions(
    xlsx_path: str,
) -> tuple[list[tuple[str, str, str, str | None, str | None]], int, int]:
    """Parse the Decisions sheet into actionable upsert rows.

    Returns:
        (rows_to_upsert, skipped_blank_count, skipped_skip_count)

    rows_to_upsert is a list of
        (competitor_id, market_code, status, reasoning_or_none, approved_by_or_none)
    tuples ready to feed into INSERT OR REPLACE. reasoning_or_none becomes the
    competitor_markets.notes column; approved_by is captured for the audit-row
    JSON but not stored on the upsert row itself (no column for it yet —
    Phase 2.2 may add an audit-trail column).
    """
    if not os.path.exists(xlsx_path):
        print(f"ERROR: input workbook not found at {xlsx_path}", file=sys.stderr)
        print(
            "  Did you mean to run scrapers/admin/build_competitor_xlsx.py first?",
            file=sys.stderr,
        )
        sys.exit(2)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Decisions" not in wb.sheetnames:
        print(
            f"ERROR: workbook has no 'Decisions' sheet. Sheets present: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(3)
    ws = wb["Decisions"]

    rows: list[tuple[str, str, str, str | None, str | None]] = []
    skipped_blank = 0
    skipped_skip = 0
    unknown_markets: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Expected order (matches scrapers/admin/build_competitor_xlsx.py:181-185):
        # (Market, Competitor, Algorithm Verdict, SERP, Queries,
        #  Best Rank, S1 URL, S2 Lang, S3 Pay, S4 App, S5 WikiFX,
        #  Decision, Reasoning, Approved By)
        if row is None or len(row) < 12:
            continue
        market_name = (row[0] or "").strip() if isinstance(row[0], str) else ""
        competitor_id = (row[1] or "").strip() if isinstance(row[1], str) else ""
        # values_only=True gives 0-based tuple indexes. Column L (12 in 1-based
        # openpyxl) → index 11 here. Column M (Reasoning) → index 12.
        # Column N (Approved By) → index 13.
        decision_raw = row[11]
        reasoning = row[12] if len(row) > 12 and row[12] else None
        approved_by = row[13] if len(row) > 13 and row[13] else None

        if not market_name or not competitor_id:
            continue  # empty row

        market_code = NAME_TO_CODE.get(market_name)
        if market_code is None:
            unknown_markets.add(market_name)
            continue

        decision = (
            (decision_raw or "").strip().lower() if isinstance(decision_raw, str) else ""
        )
        if not decision:
            skipped_blank += 1
            continue
        if decision == "skip":
            skipped_skip += 1
            continue
        if decision not in VALID_STATUSES:
            # Unknown decision value (e.g. typo in spreadsheet). Treat as
            # skip-with-warning so the operator catches it on next run.
            print(
                f"  WARN: unknown Decision '{decision}' for "
                f"({competitor_id}, {market_code}); skipping",
                file=sys.stderr,
            )
            skipped_skip += 1
            continue

        # Trim free-text columns defensively (xlsx can leak trailing whitespace).
        reasoning_clean = reasoning.strip() if isinstance(reasoning, str) else reasoning
        approved_by_clean = (
            approved_by.strip() if isinstance(approved_by, str) else approved_by
        )
        rows.append(
            (competitor_id, market_code, decision, reasoning_clean, approved_by_clean)
        )

    if unknown_markets:
        print(
            f"  WARN: unknown market name(s) skipped: {sorted(unknown_markets)}",
            file=sys.stderr,
        )

    return rows, skipped_blank, skipped_skip


def main() -> int:
    args = parse_args()
    rows, skipped_blank, skipped_skip = parse_decisions(args.input)

    if not rows:
        print(
            f"No actionable rows in {args.input} "
            f"(blank: {skipped_blank}, skip: {skipped_skip})"
        )
        return 0

    # Markets touched by this spreadsheet (used by --purge).
    markets_in_sheet = sorted({r[1] for r in rows})

    # Distinct Approved By signatures captured for the audit trail (D2.1-10).
    approvers = sorted({r[4] for r in rows if r[4]})

    now_iso = datetime.now(timezone.utc).isoformat()
    summary = {
        "rows_inserted": len(rows),
        "rows_skipped_blank": skipped_blank,
        "rows_skipped_skip": skipped_skip,
        "purge_mode": args.purge,
        "markets_in_sheet": markets_in_sheet,
        "approvers": approvers,
        "input_path": os.path.relpath(args.input, _PROJECT_ROOT),
    }

    if args.dry_run:
        print("--dry-run mode: no DB writes. Plan:")
        print(json.dumps(summary, indent=2))
        for cid, mc, status, _r, _a in rows[:5]:
            print(f"  would upsert: ({cid}, {mc}, {status})")
        if len(rows) > 5:
            print(f"  ... and {len(rows) - 5} more")
        return 0

    # contextlib.closing + WAL mode (Phase 1 invariant; matches the
    # scrapers/apify_social.py:690 pattern).
    with closing(get_db()) as conn:
        try:
            if args.purge:
                for mc in markets_in_sheet:
                    conn.execute(
                        "DELETE FROM competitor_markets WHERE market_code = ?",
                        (mc,),
                    )

            for competitor_id, market_code, status, reasoning, _approved_by in rows:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO competitor_markets
                        (competitor_id, market_code, status, notes, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (competitor_id, market_code, status, reasoning, now_iso),
                )

            # Audit row (D2.1-10). One per successful import; severity='low'
            # because this is informational, not a competitor-state change.
            conn.execute(
                """
                INSERT INTO change_events
                    (competitor_id, domain, field_name, old_value, new_value,
                     severity, detected_at, market_code)
                VALUES (?, ?, ?, NULL, ?, 'low', ?, 'global')
                """,
                (
                    AUDIT_COMPETITOR_ID,
                    AUDIT_DOMAIN,
                    AUDIT_FIELD,
                    json.dumps(summary),
                    now_iso,
                ),
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"ERROR: import failed mid-flight, rolled back: {e}", file=sys.stderr)
            return 4

    print(f"Imported {len(rows)} rows from {os.path.relpath(args.input, _PROJECT_ROOT)}")
    print(
        f"  skipped blank: {skipped_blank}, skipped 'skip': {skipped_skip}, "
        f"purge: {args.purge}"
    )
    print(f"  markets touched: {', '.join(markets_in_sheet)}")
    if approvers:
        print(f"  approvers: {', '.join(approvers)}")
    print(f"  audit row written to change_events (field_name='{AUDIT_FIELD}')")
    return 0


if __name__ == "__main__":
    sys.exit(main())
